from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from math import log2
from random import randint
import seaborn as sns


def reb(a, b):
    temp = []
    const = sorted(list("QWERTYUIOPASDFGHJKLZXCVBNM"))
    for k in const:
        temp.append(k)
    for k1 in const:
        for k2 in const:
            temp.append(k1 + k2)
    return temp.index(a) + 1, temp.index(b) + 1


def draw(year, color, another_color):
    wb1 = load_workbook(f"C:/Users/rezon/Downloads/рейт/{year}.xlsx")
    sheet = wb1["Sheet 1"]
    inf = reb("ED", "CQ")
    places = ["г.Москва", "г.Санкт-Петербург", "Свердловская область",
              "Республика Татарстан", "Новосибирская область", "Томская область"]
    print(sheet.cell(row=1, column=inf[0]).value)
    print(sheet.cell(row=1, column=inf[1]).value)

    graph = plt.subplot()
    plt.grid(visible=True)
    temp = []
    another = []
    for i in range(2, 1000):
        if sheet.cell(row=i, column=inf[0]).value and \
                sheet.cell(row=i, column=inf[1]).value and \
                (sheet.cell(row=i, column=129).value in places or True):
            temp.append((sheet.cell(row=i, column=inf[0]).value, sheet.cell(row=i, column=inf[1]).value))
    temp.sort(key=lambda x: x[0])
    q1 = temp[len(temp) // 4][0]
    q3 = temp[len(temp) - len(temp) // 4][0]
    iqr = abs(q1 - q3) * 1.5
    index = 0
    while index < len(temp):
        if temp[index][0] > q3 + iqr or temp[index][0] < q1 - iqr:
            another.append(temp[index])
            del temp[index]
        else:
            index += 1
    temp.sort(key=lambda x: x[1])
    q1 = temp[len(temp) // 4][1]
    q3 = temp[len(temp) - len(temp) // 4][1]
    iqr = abs(q1 - q3) * 1.5
    index = 0
    while index < len(temp):
        if temp[index][1] > q3 + iqr or temp[index][1] < q1 - iqr:
            another.append(temp[index])
            del temp[index]
        else:
            index += 1
    x = np.array([i[0] for i in temp])
    y = np.array([i[1] for i in temp])
    z = np.polyfit(x, y, 1)
    p = np.poly1d(z)

    plt.plot(x, p(x))

    for i in temp:
        graph.scatter(i[0], i[1], color=color)
    for i in another:
        graph.scatter(i[0], i[1], color=another_color)
    plt.xlabel(sheet.cell(row=1, column=inf[0]).value)
    plt.ylabel(sheet.cell(row=1, column=inf[1]).value)


# draw(2015, "blue", "red")
draw(2020, "blue", "red")
# draw(2021, "blue", "red")
# draw(2022, "blue", "red")
plt.show()
